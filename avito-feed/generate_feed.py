"""Собирает fleet.xlsx для Авито Автозагрузки из всех active JSON-записей в data/."""
import json
import glob
import openpyxl

HEADERS = ['Уникальный идентификатор объявления', 'Начало размещения', 'Окончание размещения', 'Способ размещения', 'Услуга продвижения', 'Номер объявления на Авито', 'Контактное лицо', 'Номер телефона', 'Способ связи', 'Описание объявления', 'Категория', 'Цена', 'Зоны показа', 'Названия фото', 'Ссылки на фото', 'Ссылка на видео', 'Адрес', 'Широта', 'Долгота', 'Название объявления', 'Адрес стоянки', 'Марка', 'Тип кузова', 'Модель', 'Колёсная формула  [Грузовики]', 'Тип двигателя [Грузовики]', 'Мощность', 'Коробка передач', 'Интернет звонки', 'Устройства для приёма звонков', 'Вид техники', 'Валюта', 'НДС включён', 'Утильсбор включён', 'Цена в валюте', 'Доступность', 'Скидка за лизинг', 'Скидка при покупке от двух единиц', 'Скидка от дилера', 'Доставка', 'Установка дополнительного оборудования', 'Официальная гарантия', 'Дополнительные условия гарантии', 'Подарки', 'URL видеофайла', 'Состояние', 'Пробег', 'ПТС или ПСМ', 'Моточасы', 'У шасси и кузова одинаковая марка?', 'Тип надстройки', 'Объём кузова', 'Длина борта', 'Ширина борта', 'Наименование кузовостроителя', 'Марка КМУ', 'Модель КМУ', 'VIN, номер кузова или SN', 'Год выпуска', 'Грузоподъёмность в кг', 'Разрешённая максимальная масса', 'Объём двигателя', 'Экологический класс']

RAW_BASE = "https://raw.githubusercontent.com/aj123123-hub/belkar-avito-feed/main/avito-feed/"

def build_row(rec):
    photo_urls = " | ".join(RAW_BASE + p for p in rec.get("photos", []))
    row = {
        'Уникальный идентификатор объявления': rec["id"],
        'Контактное лицо': rec.get("manager_name", ""),
        'Номер телефона': rec.get("phone", ""),
        'Способ связи': 'По телефону и в сообщениях',
        'Описание объявления': rec["description"],
        'Категория': 'Грузовики и спецтехника',
        'Цена': rec["price"],
        'Ссылки на фото': photo_urls,
        'Адрес': rec["address"],
        'Марка': rec["make"],
        'Тип кузова': rec["body_type"],
        'Модель': rec["model"],
        'Колёсная формула  [Грузовики]': rec["wheel_formula"],
        'Тип двигателя [Грузовики]': rec["engine_type"],
        'Мощность': rec["power_hp"],
        'Коробка передач': rec["transmission"],
        'Услуга продвижения': rec.get("promotion", ""),
        'Вид техники': 'Грузовики',
        'Валюта': rec["currency"],
        'НДС включён': 'Да' if rec.get("vat_included") else 'Нет',
        'Доступность': rec["availability"],
        'Состояние': 'С пробегом',
        'Пробег': rec["mileage_km"],
        'ПТС или ПСМ': rec["pts_or_psm"],
        'У шасси и кузова одинаковая марка?': 'Да' if rec.get("chassis_body_same_brand") else 'Нет',
        'VIN, номер кузова или SN': rec["vin"],
        'Год выпуска': rec["year"],
        'Объём кузова': rec.get("body_volume_m3", ""),
        'Экологический класс': rec.get("emission_class", ""),
    }
    return [row.get(h, "") for h in HEADERS]

def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Объявления"
    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.number_format = "@"

    count = 0
    r = 2
    for path in sorted(glob.glob("data/*.json")):
        with open(path, encoding="utf-8") as f:
            rec = json.load(f)
        if rec.get("status") != "active":
            continue
        values = build_row(rec)
        for c, v in enumerate(values, start=1):
            # требование Авито: все ячейки должны быть текстовым форматом
            cell = ws.cell(row=r, column=c, value=str(v) if v != "" else "")
            cell.number_format = "@"
        r += 1
        count += 1

    for col in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 40)

    wb.save("fleet.xlsx")
    print(f"fleet.xlsx собран: {count} активных объявлений (все ячейки — текстовый формат)")

if __name__ == "__main__":
    main()
